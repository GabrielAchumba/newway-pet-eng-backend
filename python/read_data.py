from openpyxl import load_workbook
from datetime import datetime, timedelta
from math import pow

def get_end_of_row(start_row, start_col, sheet):
    """
    Get the number of rows from the start row in a specified column until an empty cell is encountered.

    Parameters:
        start_row (int): The starting row (1-based index).
        start_col (int): The starting column (1-based index).
        sheet_name (str): The name of the worksheet.
        workbook_path (str): The path to the Excel workbook.

    Returns:
        int: The number of non-empty rows starting from start_row.
    """

    # Initialize the row count
    nrow = 0
    
    # Loop through rows until an empty cell is encountered
    while True:
        cell_value = sheet.cell(row=start_row + nrow, column=start_col).value
        if cell_value is None or str(cell_value).strip() == "":
            break
        nrow += 1
    
    return nrow


def get_end_of_col(start_row, start_col, sheet):
    """
    Get the last non-empty column in a row starting from a specific column.

    Parameters:
        start_row (int): The starting row (1-based index).
        start_col (int): The starting column (1-based index).
        sheet_name (str): The name of the worksheet.
        workbook_path (str): The path to the Excel workbook.

    Returns:
        int: The index of the last non-empty column in the specified row.
    """
    # Load the workbook and the specified sheet
    # workbook = load_workbook(workbook_path, data_only=True)
    # sheet = workbook[sheet_name]
    
    # Initialize the column index
    ncol = start_col
    
    # Loop through columns until an empty cell is encountered
    while True:
        cell_value = sheet.cell(row=start_row, column=ncol).value
        if cell_value is None or str(cell_value).strip() == "":
            break
        ncol += 1
    
    return ncol - 1  # Subtract 1 because the last increment was for an empty cell


def read_rows_data(start_row, start_col, end_row, sheet):
    """
    Reads data from a specified column and row range in an Excel sheet.

    Parameters:
        start_row (int): The starting row (1-based index).
        start_col (int): The starting column (1-based index).
        end_row (int): The last row to read (1-based index).
        sheet_name (str): The name of the worksheet.
        workbook_path (str): The path to the Excel workbook.

    Returns:
        list: A list containing the values from the specified range.
    """
    # Load the workbook and the specified sheet
    # workbook = load_workbook(workbook_path, data_only=True)
    # sheet = workbook[sheet_name]
    
    # Initialize a list to store the values
    data = []
    
    # Loop through the specified range and collect values
    for i in range(start_row, end_row + 1):
        cell_value = sheet.cell(row=i, column=start_col).value
        data.append(cell_value)
    
    return data

def read_columns_data(start_row, start_col, end_col, sheet):
    """
    Reads data from a specified row across multiple columns in an Excel sheet.

    Parameters:
        start_row (int): The starting row (1-based index).
        start_col (int): The starting column (1-based index).
        end_col (int): The last column to read (1-based index).
        sheet_name (str): The name of the worksheet.
        workbook_path (str): The path to the Excel workbook.

    Returns:
        list: A list containing the values from the specified row and column range.
    """
    # Load the workbook and the specified sheet
    # workbook = load_workbook(workbook_path, data_only=True)
    # sheet = workbook[sheet_name]
    
    # Initialize a list to store the values
    data = []
    
    # Loop through the specified range of columns
    for i in range(start_col, end_col + 1):
        cell_value = sheet.cell(row=start_row, column=i).value
        data.append(cell_value)
    
    return data

def get_numbers(start_col, end_col, x):
    """
    Extracts a subset of elements from a list based on start and end column indices.

    Parameters:
        start_col (int): The starting index (0-based index).
        end_col (int): The ending index (exclusive, 0-based index).
        x (list): The input list.

    Returns:
        list: A new list containing the elements from start_col to end_col - 1.
    """
    # Use Python slicing to extract the required range
    return x[start_col:end_col]

def read_columns_data(sheet, start_row, start_col, end_col):
    """Reads data from specified columns in a row."""
    return [sheet.cell(row=start_row, column=col).value for col in range(start_col, end_col + 1)]